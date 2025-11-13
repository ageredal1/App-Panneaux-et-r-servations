# app.py
import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import math

#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# PANNEAUX
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Cacul des panneaux, fait appel aux fonctions de panneaux
def principal_panneaux(reservoirs, fichier_excel):
    resultats_panneaux={}
    sections_nouvelles = {}   # Dictionnaire pour stocker les nouvelles sections

    for i,(t,r) in enumerate(reservoirs.items()):

        # Association levées et sections d'acier dans une liste de dictionnaires par lévée
        voile=associer_levees_svoile(r["levees"],r["s_voile"])

        # Ouvrir l'archive excel avec l'information des panneaux
        df = pd.read_excel("panneaux.xlsx", sheet_name=i+4)     # lecture de toutes les page à partir de la page indice 4

        # caracteristiques des panenaux
        panneaux = df.to_dict(orient="records")     #dataframe à liste de dictionnaires

        # Appeler la fonction de traitement
        section_voile_panneaux, panneaux_modifies = verification_Panneaux(
            panneaux, voile, r["levees"], r["pExt"], r["pInt"]
        )
        
        # Grouper les sections 
        sections_agroupees=agrouper_sections(section_voile_panneaux, t)
        sections_nouvelles[t]=sections_agroupees    # stocker les résultats

        
        """# Imprimer résultats
        print(f"Résultats pour {t}")
        pprint()
        print()"""
        

        # Sauvegarder les résultats de ce réservoir
        resultats_panneaux[t]=pd.DataFrame(panneaux_modifies)
    
    # Exporter à excel
    exporter_resultats_panneaux_excel(fichier_excel, reservoirs, resultats_panneaux, sections_nouvelles)
    return sections_nouvelles

# Exporter à excel
def exporter_resultats_panneaux_excel(fichier_excel_original, reservoirs, resultats_panneaux, sections_nouvelles):
    # Charger les feuilles excel
    xls = pd.ExcelFile(fichier_excel_original)
    writer = pd.ExcelWriter("Verification_panneaux_resultats_.xlsx", engine="openpyxl")     #Ecrir sur un nouveau fichier

    # Copier les 4 premières feuilles
    # for idx, sheet in enumerate(xls.sheet_names[:4]):
    # df_original = pd.read_excel(xls, sheet_name=sheet)
    # df_original.to_excel(writer, sheet_name=sheet, index=False)

    for idx, sheet in enumerate(xls.sheet_names[:4]):
        df_original = pd.read_excel(xls, sheet_name=sheet)

        # 🔹 Si c’est la feuille "Sections" (index 3 → sheet_names[3]), on ajoute les nouvelles valeurs
        if sheet.lower() == "sections" or idx == 3:
            df_modifie = df_original.copy()
            df_modifie["valeurs_nouvelles"] = None  # nueva columna vacía

            # Iterar por cada réservoir
            for t, data in sections_nouvelles.items():
                # data viene del resultado de agrouper_sections()
                # Ejemplo: {"Nom": "T1", "horizontal_ext": [...], "horizontal_int": [...], ...}
                for i in range(len(df_modifie)):
                    nom = str(df_modifie.loc[i, "nom"]).strip()
                    tipo = str(df_modifie.loc[i, "type"]).strip().lower()

                    # Coincide el réservoir y el tipo
                    if nom == str(data["Nom"]).strip():
                        if tipo in data:
                            valores = data[tipo]
                            # Convertir lista de floats en string con comas
                            if isinstance(valores, list):
                                df_modifie.loc[i, "valeurs_nouvelles"] = ",".join(map(str, valores))
                            else:
                                df_modifie.loc[i, "valeurs_nouvelles"] = str(valores)

            df_modifie.to_excel(writer, sheet_name=sheet, index=False)
        else:
            # Copiar sin cambios
            df_original.to_excel(writer, sheet_name=sheet, index=False)

    # Inserer les résultats à partir de la feuille 5
    for i, (t, df_out) in enumerate(resultats_panneaux.items()):
        sheet_name = xls.sheet_names[i + 4] if len(xls.sheet_names) > i + 4 else f"Resultats_{t}"
        df_out.to_excel(writer, sheet_name=sheet_name, index=False)

    # Sauvegarder
    writer.close()
    print("\n✅ Archivo 'Verification_panneaux_resultats_.xlsx' generado correctamente.")


# Fonction lecture des données des réservoirs
def donnees_Reservoirs(fichier_excel):
    """
    Charge les informations de plusieurs réservoirs depuis un fichier Excel avec trois feuilles :
      - Géométrie : colonnes ['nom', 'dr', 'ev']
      - Lévées : colonnes ['nom', 'levees'] (valeurs séparées par des virgules)
      - Sections acier : colonnes ['nom', 'type', 'valeurs'] (valeurs séparées par des virgules)
      
    Retourne un dictionnaire avec la structure :
    {
        'T1': {
            "dr": dr,
            "ev": ev,
            "h": h,
            "pExt": pExt,
            "pInt": pInt,
            "levees": levees,
            "s_voile": s_voile,
            "type": types ['Horizontal extérieur', 'Horizontal intérieur', ...]
        },
        'T2': {...},
        ...
    }
    """
    # -----------------------------
    # Lecture des feuilles Excel
    # -----------------------------
    geom = pd.read_excel(fichier_excel, sheet_name=1)
    levees_df = pd.read_excel(fichier_excel, sheet_name=2)
    s_voile_df = pd.read_excel(fichier_excel, sheet_name=3)

    donnees_reservoirs = {}

    # -----------------------------
    # Parcours de chaque réservoir
    # -----------------------------
    for t in geom["nom"]:
        # Géométrie
        dr = float(geom.loc[geom["nom"] == t, 'dr'].iloc[0])    # Diamètre du réservoir 
        ev = float(geom.loc[geom["nom"] == t, 'ev'].iloc[0])   # Epaisseur du voile 
        h = float(geom.loc[geom["nom"] == t, 'h'].iloc[0])
        pExt=math.pi*(dr+ev*2-3*2)      # Perimètre de calcul pour les panneaux ext
        pInt=math.pi*(dr+6*2)           # Perimètre de calcul pour les panneaux int


        # Lévées (chaque cellule peut contenir une ou plusieurs valeurs séparées par des virgules)
        levees_temp = levees_df[levees_df["nom"] == t]["levees"]
        levees = []
        for val in levees_temp:
            try:
                levees.append([int(float(x)) for x in str(val).split(',') if x.strip()])
            except ValueError:
                levees.append([])

        # Sections d'acier
        s_voile_temp = s_voile_df[s_voile_df["nom"] == t]
        s_voile = []
        types = []
        for _, row in s_voile_temp.iterrows():
            liste = [float(x) for x in str(row["valeurs"]).split(',')]
            s_voile.append(liste)
            types.append(row["type"])

        # Stockage des données du réservoir dans le dictionnaire
        donnees_reservoirs[t] = {
            "dr": dr,
            "ev": ev,
            "h": h,
            "pExt": pExt,
            "pInt": pInt,
            "levees": levees,
            "s_voile": s_voile,
            "tipos": types
        }

    return donnees_reservoirs


# Fonction principale de calcul
def verification_Panneaux(panneaux, voile, levees, perimetreExt, perimetreInt):
    voile_panneaux=[]       # Initialiser une liste de dictionnaires pour ajouter la section de chaque panneaux et la levée associer 

    # ------------------------------
    # Verification
    # ------------------------------
    for panneau in panneaux:
        # Verification longueur
        if panneau["position"]=="ext":
            if verif_longueur_panneaux(panneau["longueur"], panneau["recouvrement"], perimetreExt, panneau["u"]):
                panneau["Verification longeur"]="OK"
            else:
                panneau["Verification longeur"]="Trop court"
        elif panneau["position"]=="int":
            if verif_longueur_panneaux(panneau["longueur"], panneau["recouvrement"], perimetreInt, panneau["u"]):
                panneau["Verification longeur"]="OK"
            else:
                panneau["Verification longeur"]="Trop court"
        else:
            panneau["Verification longeur"]="Verifier position"

        # Calcul des sections
        sx1=round(calcul_section(panneau["ex1"], panneau["dx1"], panneau["n1"]),2)       #Section horizontal premier metre
        sx2 = round(calcul_section(panneau["ex2"], panneau["dx2"], panneau["n2"]),2) if panneau["n2"] > 0 else None       #Section horizontal deuxiem metre
        sy=round(calcul_section(panneau["ey"], panneau["dy"], 1),2)                      #Section vertical
        panneau["sx1"]=sx1
        panneau["sx2"] = sx2 if sx2 is not None else None
        panneau["sy"]=sy


        # Verification des sections
        # extérieur
        if panneau["position"]=="ext":
            # sx1
            if panneau["sx1"]>=voile[panneau["levee"]-1]["horizontal_ext"][0]:
                panneau["Verif. sx1"]="OK"
            else:
                panneau["Verif. sx1"]="Augmenter la section"
            # sx2
            if len(levees[panneau["levee"]-1])>1:
                if panneau["sx2"]>=voile[panneau["levee"]-1]["horizontal_ext"][1]:
                    panneau["Verif. sx2"]="OK"
                else:
                    panneau["Verif. sx2"]="Augmenter la section"
            # sy
            if panneau["sy"]>=voile[panneau["levee"]-1]["vertical_ext"]:
                panneau["Verif. sy"]="OK"
            else:
                panneau["Verif. sy"]="Augmenter la section"
            

        # intérieur
        if panneau["position"]=="int":
            # sx1
            if panneau["sx1"]>=voile[panneau["levee"]-1]["horizontal_int"][0]:
                panneau["Verif. sx1"]="OK"
            else:
                panneau["Verif. sx1"]="Augmenter la section"
            # sx2
            if len(levees[panneau["levee"]-1])>1:
                if panneau["sx2"]>=voile[panneau["levee"]-1]["horizontal_int"][1]:
                    panneau["Verif. sx2"]="OK"
                else:
                    panneau["Verif. sx2"]="Augmenter la section"
            # sy
            if panneau["sy"]>=voile[panneau["levee"]-1]["vertical_int"]:
                panneau["Verif. sy"]="OK"
            else:
                panneau["Verif. sy"]="Augmenter la section"

        # Dictionnaire du panneau
        voile_panneau = {
        "levee": panneau["levee"],
        "horizontal_ext": [sx1] + ([sx2] if sx2 is not None else []) if panneau["position"]=="ext" else None,
        "horizontal_int": [sx1] + ([sx2] if sx2 is not None else []) if panneau["position"]=="int" else None,
        "vertical_ext": sy if panneau["position"]=="ext" else 0,
        "vertical_int": sy if panneau["position"]=="int" else 0
        }
            
        voile_panneaux.append(voile_panneau)    #Ajouter le dictionnaire du panneau à la liste des panneaux


    # Sommer les sections des panneaux (dans le cas s'il y a des doubles panneaux)
   

    # Convertir à dataframe
    df_voile = pd.DataFrame(voile_panneaux)

    # Grouper par levee en additionnant des listes par index
    result = {}     # Dictionnaire auxiliaire

    for _, row in df_voile.iterrows():      #iterer par chaque ligne du DataFrame
        lev = row["levee"]

        # Initialiser le dictionnaire si c'est la première fois de la levée
        if lev not in result:
            result[lev] = {
                "horizontal_ext": [0]*len(row["horizontal_ext"]) if row["horizontal_ext"] else [],  #liste de zéros
                "horizontal_int": [0]*len(row["horizontal_int"]) if row["horizontal_int"] else [],  #liste de zéros
                "vertical_ext": 0,  #Valeur 0
                "vertical_int": 0   #Valeur 0
            }

        # Somme horizontal_ext
        if row["horizontal_ext"]:
            # Ajuster la taille de la liste cumulée si elle est inférieure à la ligne actuelle
            if len(result[lev]["horizontal_ext"]) < len(row["horizontal_ext"]):
                result[lev]["horizontal_ext"] += [0]*(len(row["horizontal_ext"]) - len(result[lev]["horizontal_ext"]))
            result[lev]["horizontal_ext"] = [x + y for x, y in zip(result[lev]["horizontal_ext"], row["horizontal_ext"])]

        # Somme horizontal_int
        if row["horizontal_int"]:
            if len(result[lev]["horizontal_int"]) < len(row["horizontal_int"]):
                result[lev]["horizontal_int"] += [0]*(len(row["horizontal_int"]) - len(result[lev]["horizontal_int"]))
            result[lev]["horizontal_int"] = [x + y for x, y in zip(result[lev]["horizontal_int"], row["horizontal_int"])]

        # Somme vertical
        if row["vertical_ext"] is not None:
            result[lev]["vertical_ext"] += row["vertical_ext"]
        if row["vertical_int"] is not None:
            result[lev]["vertical_int"] += row["vertical_int"]

    # -----------------------------
    # Convertir a liste de diccionarios
    # -----------------------------
    section_voile_panneaux = [{"levee": k, **v} for k, v in result.items()]
    return section_voile_panneaux, panneaux


# Fonction pour calculer la section d'acier avec l'espacement, le diamètre de la barre et le nombre des barres
def calcul_section(e,d,n):
    A=math.pi*((d/10)**2)/4
    return A*(1000/e)*n if e!=0 else 0


# Fonction pour verifier la longuer des panneaux pour couvrir le perimètre de l'ouvrage
def verif_longueur_panneaux(longueur,recouvrement,perimetre,u):
    perimetrePanneau=(longueur-recouvrement)*u
    return perimetrePanneau>perimetre


# Fontion pour associer les levées et les sections d'acier
def associer_levees_svoile(levees, s_voile):
    associations = []

    # Position initial des listes horizontales
    pos_horiz = [0, 0]

    for i, levee in enumerate(levees):
        n = len(levee)  # nombre des éléments par bloc

        # horizontal_ext
        horizontal_ext = s_voile[0][pos_horiz[0]:pos_horiz[0]+n] if len(s_voile) > 0 else []
        # horizontal_int
        horizontal_int = s_voile[1][pos_horiz[1]:pos_horiz[1]+n] if len(s_voile) > 1 else []

        pos_horiz[0] += n
        pos_horiz[1] += n

        # vertical_ext
        vertical_ext = s_voile[2][i] if len(s_voile) > 2 and i < len(s_voile[2]) else []
        # vertical_int
        vertical_int = s_voile[3][i] if len(s_voile) > 3 and i < len(s_voile[3]) else []

        association = {
            "levee": levee,
            "horizontal_ext": horizontal_ext,
            "horizontal_int": horizontal_int,
            "vertical_ext": vertical_ext,
            "vertical_int": vertical_int
        }
        associations.append(association)

    return associations

# Fonction pour agrouper les sections du voile
def agrouper_sections(sections_voile, t):
    hext=[]
    hint=[]
    vext=[]
    vint=[]

    for l in sections_voile:
        hext=hext+(l["horizontal_ext"])
        hint=hint+(l["horizontal_int"])
        vext.append(l["vertical_ext"])
        vint.append(l["vertical_int"])

    sections={"Nom":t,
              "horizontal_ext":hext,
              "horizontal_int":hint,
              "vertical_ext":vext,
              "vertical_int":vint}
    
    return sections 


#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# RESERVATIONS
#------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# Princial des réservations
def principal_reservations(reservations,sec_panneaux, reservoirs):
    for i,(t,r) in enumerate(reservations.items()):
        levees=reservoirs[t]["levees"]
        for ri in r:
            if pd.isna(ri["diametre"]):
                # rectangulaire
                # barres horizontales
                nx, lgx, scx, sx= calcul_reservation(reservoirs[t]["h"],
                                                     ri["hauteur_axe"],
                                                     ri["hauteur"],
                                                     sec_panneaux[t]["horizontal_ext"],
                                                     "h",
                                                     ri["hauteur"],
                                                     ri["dbarre_x"],
                                                     ri["largeur"],
                                                     levees)

                # barres verticales
                ny, lgy, scy, sy= calcul_reservation(reservoirs[t]["h"],
                                                     ri["hauteur_axe"],
                                                     ri["hauteur"],
                                                     sec_panneaux[t]["vertical_ext"],
                                                     "v",
                                                     ri["largeur"],
                                                     ri["dbarre_y"],
                                                     ri["hauteur"],
                                                     levees)
                
            else:
                # circualire
                # barres horizontales
                nx, lgx, scx, sx= calcul_reservation(reservoirs[t]["h"],
                                                     ri["hauteur_axe"],
                                                     ri["diametre"],
                                                     sec_panneaux[t]["horizontal_ext"],
                                                     "h",
                                                     ri["diametre"],
                                                     ri["dbarre_x"],
                                                     ri["diametre"],
                                                     levees)

                #barres verticales
                ny, lgy, scy, sy= calcul_reservation(reservoirs[t]["h"],
                                                     ri["hauteur_axe"],
                                                     ri["diametre"],
                                                     sec_panneaux[t]["vertical_ext"],
                                                     "v",
                                                     ri["diametre"],
                                                     ri["dbarre_y"],
                                                     ri["diametre"],
                                                     levees)
            
            # Enregistrer les informations dans le dictionaire
            ri["Section x coupée (cm²/m)"]=scx
            ri["Nb barres x min par face"]=nx
            ri["Longueur x min (cm)"]=lgx
            ri["Section d'acier x mise en place (cm²/m)"]=sx

            ri["Section y coupée (cm²/m)"]=scy
            ri["Nb barres y min par face"]=ny
            ri["Longueur y min (cm)"]=lgy
            ri["Section d'acier y mise en place (cm²/m)"]=sy
    exporter_reservations(reservations)

# Lecture excel
def lire_Excel_reservations(nomFichier):
    donnees=pd.read_excel(nomFichier, sheet_name=None)   # Lecture de toutes les page {nom_page: DataFrame}
    nom_pages=list(donnees.keys())
    nombre_pages=len(nom_pages)
    
    # Dataframe à liste de dictionnaires
    reservations={nom: df.to_dict(orient="records") for nom, df in donnees.items()}

    return reservations

# Section coupée
def section_coupee(hvoile, hauteur_axe, hauteur, s_voile, sens, dim_qui_coupe, levees):     # Si c'est un cercle: largeur=hauteur=diametre
    hvoile_m=hvoile/100
    h_axe=hauteur_axe   # cm
    h_point_bas=(h_axe-hauteur/2)/100   # de cm à m
    h_point_haut=(h_axe+hauteur/2)/100  # de cm à m

    # Bornes de sécurité
    h_point_bas = max(0, h_point_bas)
    h_point_haut = min(h_point_haut, hvoile_m)
    
    if sens=="h":
        sec_point_bas=s_voile[int(h_point_bas)]
        sec_point_haut=s_voile[int(h_point_haut)]
        if sec_point_bas==sec_point_haut:
            return sec_point_bas*(dim_qui_coupe/100)
        else:
            sec_total=sec_point_bas*(int(h_point_haut)-h_point_bas)/100 + sec_point_haut*(h_point_haut-int(h_point_haut))/100
            return sec_total
    else:
        levee=ind_levee(levees, h_point_bas)
        sec_v=s_voile[levee]
        return sec_v*(dim_qui_coupe/100)    

# indice levee
def ind_levee(levees, h):   # Donne l'indice de la levée de la réservation 
    for i, (a,b) in enumerate(levees):
        if a-1<= h < b:
            return i
    
    return None

# Cacul section de la barre
def aireBarre(d):
    return math.pi*((d/10)/2)**2

# Nombre de barres à mettre en place
def nBarres(section,diametreBarre):
    n=math.ceil(section/aireBarre(diametreBarre))
    return n if n % 2 == 0 else n + 1

# longueur des barres
def long_Barres(dim1Trou,dbarre,dim2Trou):
    lg=dim1Trou+2*34*(dbarre/10)+dim2Trou/2
    return math.ceil(lg/10)*10

# Section mise en place
def sec_mise_en_place(n,db):
    return n*aireBarre(db)

# Calcul de la réservation
def calcul_reservation(hvoile, haxe, hauteur, sec_panneau, sens, dim_qui_coupee, dbarre, dim2, levees):
    sc=section_coupee(hvoile, haxe, hauteur, sec_panneau, sens, dim_qui_coupee, levees)   # section coupée
    n=nBarres(sc, dbarre)                                                                          # nombre des barres
    s=sec_mise_en_place(n, dbarre)                                                                 # section mise en place
    lg=long_Barres(dim2, dbarre, dim_qui_coupee)                                                   # Longueur des barres
    return n, lg, sc, s

# Exporter excel réservations
def exporter_reservations(reservations, nom_fichier="Renfort_reservations_avec_altimetrie.xlsx"):
    """
    Exporte toutes les réservations dans un seul fichier Excel,
    avec une feuille par réservoir (clé du dictionnaire 'reservations').
    """
   # Créer un nouveau classeur
    wb = Workbook()
    wb.remove(wb.active)  # Supprime la feuille par défaut

    # Définir style de bordure
    bordure = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Parcourir chaque réservoir
    for nom_reservoir, liste_reservations in reservations.items():
        ws = wb.create_sheet(title=nom_reservoir[:31])  # Excel limite à 31 caractères

        ligne = 1
        for reservation in liste_reservations:
            # Titre
            ws.cell(row=ligne, column=1, value=reservation["NOM"])
            ws.cell(row=ligne, column=1).font = Font(bold=True, size=12)
            ligne += 1

            # En-têtes
            headers = [
                "Sens",
                "Dimension (cm)",
                "Section coupée (cm²/m)",
                "Diamètre barre (mm)",
                "Nb barres min par face",
                "Longueur min (cm)",
                "Section mise en place (cm²/m)"
            ]

            for col, h in enumerate(headers, start=1):
                c = ws.cell(row=ligne, column=col, value=h)
                c.font = Font(bold=True)
                c.border = bordure
                c.alignment = Alignment(horizontal="center", vertical="center")

            ligne += 1

            # --- Déterminer la dimension principale selon la forme ---
            # Si diametre existe => circulaire, sinon rectangulaire
            diam = reservation.get("diametre", None)
            if pd.notna(diam):
                dim_x = dim_y = f"ø{diam}"
            else:
                dim_x = reservation.get("largeur", "")
                dim_y = reservation.get("hauteur", "")

            # Lignes de données (Horizontal et Vertical)
            try:
                lignes_table = [
                [
                    "Horizontal",
                    dim_x,  # dimension selon forme
                    round(reservation.get("Section x coupée (cm²/m)", 0), 2),
                    f"Ø{reservation.get('dbarre_x', '')}",
                    reservation.get("Nb barres x min par face", ""),
                    round(reservation.get("Longueur x min (cm)", 0), 2),
                    round(reservation.get("Section d'acier x mise en place (cm²/m)", 0), 2),
                ],
                [
                    "Vertical",
                    dim_y,  # dimension selon forme
                    round(reservation.get("Section y coupée (cm²/m)", 0), 2),
                    f"Ø{reservation.get('dbarre_y', '')}",
                    reservation.get("Nb barres y min par face", ""),
                    round(reservation.get("Longueur y min (cm)", 0), 2),
                    round(reservation.get("Section d'acier y mise en place (cm²/m)", 0), 2),
                ]
                ]
            except Exception as e:
                print(f"⚠️ Erreur dans la réservation {reservation.get('NOM')}: {e}")
                continue

            # Écrire les lignes de données
            for row_data in lignes_table:
                for col, value in enumerate(row_data, start=1):
                    c = ws.cell(row=ligne, column=col, value=value)
                    c.border = bordure
                    c.alignment = Alignment(horizontal="center", vertical="center")
                ligne += 1

            ligne += 2  # espace entre réservations

        # Ajustement automatique de la largeur des colonnes
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(nom_fichier)
    print(f"✅ Exportation réussie : {nom_fichier} avec {len(reservations)} feuilles.")

# ----------------------------
# Interface Streamlit
# ----------------------------
st.set_page_config(page_title="Vérification des panneaux et réservations")

st.title("✅ Vérification des panneaux et calcul du renfort des réservations")
st.write("Téléchargez vos fichiers Excel et cliquez sur 'Lancer le calcul'")

# Charger fichiers Excel
file_panneaux = st.file_uploader("📁 Panneaux.xlsx", type=["xlsx"])
file_reservations = st.file_uploader("📁 Reservations_avec_altimetrie.xlsx", type=["xlsx"])

if file_panneaux and file_reservations:
    if st.button("Lancer le calcul"):

        # Lire fichiers Excel
        st.info("📥 Lecture des fichiers...")
        reservoirs = donnees_Reservoirs(file_panneaux)
        sections_voile = principal_panneaux(reservoirs, file_panneaux)

        Reservations = lire_Excel_reservations(file_reservations)
        principal_reservations(Reservations, sections_voile, reservoirs)

        # ----------------------------
        # Préparer fichiers Excel pour téléchargement
        # ----------------------------
        st.success("✅ Calcul terminé ! Téléchargez vos résultats ci-dessous.")

        # 1. Verification_panneaux_resultats_.xlsx
        with BytesIO() as buffer_panneaux:
            # Votre fonction exporter_resultats_panneaux_excel adaptée pour BytesIO
            exporter_resultats_panneaux_excel(file_panneaux, reservoirs, sections_voile, sections_voile, output_buffer=buffer_panneaux)
            buffer_panneaux.seek(0)
            st.download_button(
                label="📥 Télécharger Verification_panneaux_resultats_.xlsx",
                data=buffer_panneaux,
                file_name="Verification_panneaux_resultats_.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        # 2. Renfort_reservations_avec_altimetrie.xlsx
        with BytesIO() as buffer_reservations:
            exporter_reservations(Reservations, nom_fichier=buffer_reservations)
            buffer_reservations.seek(0)
            st.download_button(
                label="📥 Télécharger Renfort_reservations_avec_altimetrie.xlsx",
                data=buffer_reservations,
                file_name="Renfort_reservations_avec_altimetrie.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )